import csv, hashlib, io, zipfile
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from xml.etree.ElementTree import ParseError

from flask import g

from extensions import db
from models.client import Client
from models.consumer_unit import ConsumerUnit
from models.import_preview import ImportPreview
from models.log_entry import LogEntry
from models.plant import Plant

MAX_BYTES, MAX_ROWS, TTL_MINUTES, MAX_COLUMNS, MAX_CELL_CHARS = 10 * 1024 * 1024, 10_000, 20, 80, 2_000
SHEETS = {'Clientes': 'clientes', 'UCs': 'ucs', 'Usinas': 'usinas'}
REQUIRED = {'clientes': {'nome', 'cpf', 'email'}, 'ucs': {'clienteCpf', 'codigo'}, 'usinas': {'nome', 'uc', 'kwPico'}}


def criar_preview(file_storage, tipo_csv: str | None) -> dict:
    purge_expirados(empresa_id=g.current_empresa_id)
    content = file_storage.read(MAX_BYTES + 1)
    if not content or len(content) > MAX_BYTES:
        raise ValueError('Arquivo vazio ou maior que 10 MB.')
    filename = (file_storage.filename or '').lower()
    arquivo_hash = hashlib.sha256(content).hexdigest()
    try:
        rows = _read_file(content, filename, tipo_csv)
    except ValueError:
        _audit_preview_rejeitado(arquivo_hash)
        raise
    plan, errors = _validate(rows)
    preview = ImportPreview(
        empresa_id=g.current_empresa_id, created_by_id=g.current_user.id,
        arquivo_hash=arquivo_hash, plano=plan,
        status='pronto' if not errors else 'invalido', expires_at=datetime.utcnow() + timedelta(minutes=TTL_MINUTES),
    )
    db.session.add(preview)
    db.session.flush()
    _audit('import_preview_created', preview, 'pronto' if not errors else 'invalido')
    db.session.commit()
    return {'previewId': preview.id, 'expiraEm': preview.expires_at.isoformat(), 'contagens': {k: len(v) for k,v in plan.items()}, 'erros': errors}


def confirmar(preview_id: int) -> dict | None:
    purge_expirados(empresa_id=g.current_empresa_id)
    preview = ImportPreview.query.filter_by(id=preview_id, empresa_id=g.current_empresa_id, created_by_id=g.current_user.id).first()
    if not preview:
        _audit_confirmacao_rejeitada()
        return None
    if preview.status != 'pronto' or preview.expires_at < datetime.utcnow():
        _audit('import_confirm_rejected', preview, 'indisponivel', nivel='warning')
        db.session.commit()
        raise ValueError('Preview indisponivel, expirado ou ja utilizado.')
    plan = preview.plano
    try:
        clients = {}
        for row in plan['clientes']:
            cpf = _digits(row['cpf'])
            if Client.query.filter(Client.empresa_id == g.current_empresa_id, Client.cpf == cpf).first():
                raise ValueError('Cliente duplicado no banco.')
            client = Client(empresa_id=g.current_empresa_id, nome=row['nome'], cpf=cpf, email=row['email'], telefone=row.get('telefone'), concessionaria=row.get('concessionaria') or 'Copel')
            db.session.add(client); db.session.flush(); clients[cpf] = client
        for row in plan['ucs']:
            cpf = _digits(row['clienteCpf']); client = clients.get(cpf)
            if not client: raise ValueError('UC referencia cliente ausente no mesmo arquivo.')
            if ConsumerUnit.query.filter(ConsumerUnit.empresa_id == g.current_empresa_id, ConsumerUnit.client_id == client.id, ConsumerUnit.codigo == row['codigo']).first(): raise ValueError('UC duplicada no banco.')
            db.session.add(ConsumerUnit(empresa_id=g.current_empresa_id, client_id=client.id, codigo=row['codigo'], consumo=_number(row.get('consumo')), concessionaria=row.get('concessionaria')))
        for row in plan['usinas']:
            if Plant.query.filter(Plant.empresa_id == g.current_empresa_id, Plant.nome == row['nome'], Plant.uc == row['uc']).first(): raise ValueError('Usina duplicada no banco.')
            db.session.add(Plant(empresa_id=g.current_empresa_id, nome=row['nome'], uc=row['uc'], kw_pico=_number(row['kwPico']), concessionaria=row.get('concessionaria')))
        preview.status = 'consumido'
        _audit('import_confirmed', preview, 'sucesso')
        db.session.commit()
    except Exception:
        db.session.rollback()
        _audit('import_confirm_failed', preview, 'falha', nivel='warning')
        db.session.commit()
        raise
    return {'clientes': len(plan['clientes']), 'ucs': len(plan['ucs']), 'usinas': len(plan['usinas'])}


def purge_expirados(empresa_id: int | None = None) -> int:
    """Remove planos expirados; sem empresa é exclusivo para o job operacional global."""
    query = ImportPreview.query.filter(ImportPreview.expires_at < datetime.utcnow())
    if empresa_id is not None:
        query = query.filter(ImportPreview.empresa_id == empresa_id)
    count = query.delete(synchronize_session=False)
    db.session.commit()
    # A bulk delete bypasses ORM bookkeeping; avoid exposing stale PII from the
    # current request's identity map after cleanup.
    db.session.expire_all()
    return count


def _unsafe_cell(value) -> bool:
    return isinstance(value, str) and value.lstrip().startswith(('=', '+', '-', '@'))


def _read_file(content, filename, tipo_csv):
    if filename.endswith('.xlsm') or not (filename.endswith('.csv') or filename.endswith('.xlsx')): raise ValueError('Use CSV UTF-8 ou XLSX.')
    if filename.endswith('.csv'):
        if tipo_csv not in REQUIRED: raise ValueError('CSV requer tipo: clientes, ucs ou usinas.')
        try: data = content.decode('utf-8-sig')
        except UnicodeDecodeError as exc: raise ValueError('CSV deve usar UTF-8.') from exc
        reader = csv.DictReader(io.StringIO(data))
        if not reader.fieldnames or len(reader.fieldnames) > MAX_COLUMNS: raise ValueError('CSV excede limite de colunas.')
        result=[]
        for index,row in enumerate(reader, 1):
            if index > MAX_ROWS: raise ValueError('Máximo de 10 mil linhas.')
            if any(_unsafe_cell(v) or len(str(v or '')) > MAX_CELL_CHARS for v in row.values()): raise ValueError('Fórmulas ou célula excessiva não são permitidas.')
            result.append(row)
        return {tipo_csv: result}
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as exc:
        raise RuntimeError('Suporte XLSX indisponivel.') from exc
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            if len(archive.infolist()) > 200 or sum(i.file_size for i in archive.infolist()) > MAX_BYTES * 5: raise ValueError('XLSX suspeito ou excede limite descompactado.')
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    except (zipfile.BadZipFile, InvalidFileException, ParseError, OSError, KeyError, ValueError) as exc:
        raise ValueError('XLSX inválido.') from exc
    if set(workbook.sheetnames) - set(SHEETS): raise ValueError('XLSX contém aba não permitida.')
    result, total_rows = {}, 0
    try:
        for sheet, kind in SHEETS.items():
            if sheet not in workbook.sheetnames: result[kind] = []; continue
            iterator = workbook[sheet].iter_rows(values_only=False)
            header_row = next(iterator, None)
            if not header_row: result[kind] = []; continue
            if len(header_row) > MAX_COLUMNS: raise ValueError('XLSX excede limite de colunas.')
            headers = [str(c.value or '').strip() for c in header_row]
            result[kind] = []
            for row in iterator:
                total_rows += 1
                if total_rows > MAX_ROWS: raise ValueError('Máximo de 10 mil linhas.')
                if any(_unsafe_cell(c.value) or len(str(c.value or '')) > MAX_CELL_CHARS for c in row): raise ValueError('Fórmulas ou célula excessiva não são permitidas.')
                result[kind].append({headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)})
    except (ParseError, OSError, KeyError, zipfile.BadZipFile, InvalidFileException) as exc:
        raise ValueError('XLSX inválido.') from exc
    finally:
        workbook.close()
    return result


def _validate(rows):
    if sum(len(v) for v in rows.values()) > MAX_ROWS: raise ValueError('Máximo de 10 mil linhas.')
    errors, plan = [], {k: [] for k in REQUIRED}
    seen = set()
    for kind, entries in rows.items():
        for index, raw in enumerate(entries, 2):
            clean = {k: str(v).strip() if v is not None else '' for k,v in raw.items()}
            missing = [f for f in REQUIRED[kind] if not clean.get(f)]
            if missing: errors.append({'tipo': kind, 'linha': index, 'erro': f'Campos obrigatórios: {", ".join(missing)}'}); continue
            if kind == 'clientes':
                clean['cpf'] = _digits(clean['cpf']); key=(kind,clean['cpf'])
                if len(clean['cpf']) != 11 or '@' not in clean['email']: errors.append({'tipo':kind,'linha':index,'erro':'CPF ou email inválido.'}); continue
            elif kind == 'ucs': key=(kind,_digits(clean['clienteCpf']),clean['codigo'])
            else:
                key=(kind,clean['nome'],clean['uc'])
                try: _number(clean['kwPico'])
                except ValueError: errors.append({'tipo':kind,'linha':index,'erro':'kwPico inválido.'}); continue
            if key in seen: errors.append({'tipo':kind,'linha':index,'erro':'Duplicata no arquivo.'}); continue
            seen.add(key); plan[kind].append(clean)
    return plan, errors

def _digits(value): return ''.join(c for c in str(value) if c.isdigit())
def _number(value):
    try: return Decimal(str(value).replace(',', '.'))
    except InvalidOperation as exc: raise ValueError('Número inválido.') from exc


def _counts(plan: dict) -> dict:
    return {kind: len(rows) for kind, rows in plan.items()}


def _audit(acao: str, preview: ImportPreview, resultado: str, *, nivel: str = 'info') -> None:
    """Auditoria imutável e deliberadamente sem conteúdo da planilha."""
    db.session.add(LogEntry(
        empresa_id=g.current_empresa_id, nivel=nivel, acao=acao,
        entidade='ImportPreview', entidade_id=preview.id,
        mensagem='Importação em massa processada',
        metadados={'usuarioId': g.current_user.id, 'arquivoHash': preview.arquivo_hash,
                   'contagens': _counts(preview.plano), 'resultado': resultado},
    ))


def _audit_preview_rejeitado(arquivo_hash: str) -> None:
    db.session.add(LogEntry(
        empresa_id=g.current_empresa_id, nivel='warning', acao='import_preview_rejected',
        entidade='ImportPreview', mensagem='Preview de importação rejeitado',
        metadados={'usuarioId': g.current_user.id, 'arquivoHash': arquivo_hash,
                   'contagens': {}, 'resultado': 'rejeitado'},
    ))
    db.session.commit()


def _audit_confirmacao_rejeitada() -> None:
    db.session.add(LogEntry(
        empresa_id=g.current_empresa_id, nivel='warning', acao='import_confirm_rejected',
        entidade='ImportPreview', mensagem='Confirmação de importação rejeitada',
        metadados={'usuarioId': g.current_user.id, 'contagens': {}, 'resultado': 'nao_encontrado'},
    ))
    db.session.commit()
