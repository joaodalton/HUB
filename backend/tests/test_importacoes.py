"""Preview/commit de importação: parsing, RBAC, isolamento e atomicidade."""
import io, os, sys, tempfile, unittest, zipfile
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook

_DB = tempfile.NamedTemporaryFile(suffix='.db', delete=False); _DB.close()
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from app import create_app  # noqa
from config import Config  # noqa
from extensions import db, limiter  # noqa
from models.client import Client  # noqa
from models.empresa import Empresa  # noqa
from models.import_preview import ImportPreview  # noqa
from models.log_entry import LogEntry  # noqa
from models.user import User  # noqa
from services.import_service import MAX_CELL_CHARS, MAX_COLUMNS, MAX_ROWS, _read_file, criar_preview  # noqa
from utils.auth import generate_token  # noqa
from werkzeug.datastructures import FileStorage  # noqa
try:
 from .support import IsolatedTestRuntime  # noqa
except ImportError:
 from support import IsolatedTestRuntime  # noqa

class ImportacoesTest(IsolatedTestRuntime, unittest.TestCase):
 @classmethod
 def setUpClass(cls):
  cls.prepare_test_runtime(f"sqlite:///{_DB.name.replace(chr(92), '/')}",'import-test-secret',limiter_enabled=False)
  cls.app=create_app(); cls.app.config.update(TESTING=True, RATELIMIT_ENABLED=False)
  with cls.app.app_context():
   db.create_all(); a=Empresa(nome='A',slug='import-a'); b=Empresa(nome='B',slug='import-b'); db.session.add_all([a,b]); db.session.flush()
   users=[User(empresa_id=a.id,nome='Owner',email='io@example.test',password_hash='x',role='owner'),User(empresa_id=a.id,nome='Viewer',email='iv@example.test',password_hash='x',role='viewer'),User(empresa_id=b.id,nome='Owner B',email='ib@example.test',password_hash='x',role='owner')]
   db.session.add_all(users); db.session.commit(); cls.a,cls.viewer,cls.b=[u.id for u in users]
 @classmethod
 def tearDownClass(cls):
  with cls.app.app_context(): db.session.remove(); db.drop_all(); db.engine.dispose()
  os.unlink(_DB.name)
  cls.restore_test_runtime()
 def _token(self, user):
  with self.app.app_context(): return generate_token(user)
 def _preview(self,user,body,kind='clientes'):
  return self.app.test_client().post('/api/v1/importacoes/preview',headers={'Authorization':'Bearer '+self._token(user)},data={'tipo':kind,'arquivo':(io.BytesIO(body), 'dados.csv')},content_type='multipart/form-data')
 def _commit(self,user,pid): return self.app.test_client().post(f'/api/v1/importacoes/{pid}/confirmar',headers={'Authorization':'Bearer '+self._token(user)})
 def test_csv_preview_has_no_writes_then_atomic_commit_and_replay(self):
  r=self._preview(self.a,b'nome,cpf,email\nAna,12345678901,ana@example.test\n'); self.assertEqual(r.status_code,201); pid=r.json['data']['previewId']
  with self.app.app_context(): self.assertEqual(Client.query.count(),0); self.assertEqual(ImportPreview.query.count(),1)
  self.assertEqual(self._commit(self.a,pid).status_code,200)
  with self.app.app_context():
   self.assertEqual(Client.query.count(),1)
   audit=LogEntry.query.filter_by(acao='import_confirmed', entidade_id=pid).one()
   self.assertEqual(audit.metadados['resultado'],'sucesso'); self.assertEqual(audit.metadados['usuarioId'],self.a); self.assertEqual(audit.metadados['contagens']['clientes'],1)
  self.assertEqual(self._commit(self.a,pid).status_code,409)
 def test_xlsx_valid_and_formula_rejected(self):
  wb=Workbook(); ws=wb.active; ws.title='Clientes'; ws.append(['nome','cpf','email']); ws.append(['Bia','12345678902','bia@example.test']); out=io.BytesIO(); wb.save(out)
  r=self.app.test_client().post('/api/v1/importacoes/preview',headers={'Authorization':'Bearer '+self._token(self.a)},data={'arquivo':(io.BytesIO(out.getvalue()),'dados.xlsx')},content_type='multipart/form-data'); self.assertEqual(r.status_code,201)
  wb=Workbook(); ws=wb.active; ws.title='Clientes'; ws.append(['nome','cpf','email']); ws.append(['=NOW()','12345678903','x@example.test']); out=io.BytesIO(); wb.save(out)
  bad=self.app.test_client().post('/api/v1/importacoes/preview',headers={'Authorization':'Bearer '+self._token(self.a)},data={'arquivo':(io.BytesIO(out.getvalue()),'formula.xlsx')},content_type='multipart/form-data'); self.assertEqual(bad.status_code,400)
  xlsm=self.app.test_client().post('/api/v1/importacoes/preview',headers={'Authorization':'Bearer '+self._token(self.a)},data={'arquivo':(io.BytesIO(b'not-a-workbook'),'macro.xlsm')},content_type='multipart/form-data'); self.assertEqual(xlsm.status_code,400)
 def test_parser_rejects_all_formula_prefixes_and_resource_limits(self):
  for prefix in ('=', '+', '-', '@'):
   with self.assertRaises(ValueError): _read_file(f'nome,cpf,email\n{prefix}X,12345678901,x@example.test\n'.encode(), 'dados.csv', 'clientes')
   wb=Workbook(); ws=wb.active; ws.title='Clientes'; ws.append(['nome','cpf','email']); ws.append([prefix+'X','12345678901','x@example.test']); out=io.BytesIO(); wb.save(out)
   with self.assertRaises(ValueError): _read_file(out.getvalue(), 'dados.xlsx', None)
  with self.assertRaises(ValueError): _read_file((','.join(f'c{i}' for i in range(MAX_COLUMNS + 1)) + '\n').encode(), 'wide.csv', 'clientes')
  with self.assertRaises(ValueError): _read_file(f'nome,cpf,email\n{"x" * (MAX_CELL_CHARS + 1)},12345678901,x@example.test\n'.encode(), 'large.csv', 'clientes')
  wb=Workbook(); ws=wb.active; ws.title='Clientes'; ws.append(['nome','cpf','email']); ws.cell(row=MAX_ROWS + 2, column=1, value='sparse'); out=io.BytesIO(); wb.save(out)
  with self.assertRaises(ValueError): _read_file(out.getvalue(), 'sparse.xlsx', None)
 def test_invalid_xlsx_returns_400_and_expired_preview_is_purged_before_preview(self):
  invalid=self.app.test_client().post('/api/v1/importacoes/preview',headers={'Authorization':'Bearer '+self._token(self.a)},data={'arquivo':(io.BytesIO(b'not an xlsx'),'invalido.xlsx')},content_type='multipart/form-data'); self.assertEqual(invalid.status_code,400)
  with self.app.test_request_context('/'):
   from flask import g
   g.current_empresa_id=1; g.current_user=db.session.get(User,self.a)
   first=criar_preview(FileStorage(stream=io.BytesIO(b'nome,cpf,email\nPurge,12345678907,purge@example.test\n'), filename='purge.csv'), 'clientes')['previewId']
   preview=db.session.get(ImportPreview,first); preview.expires_at=datetime.utcnow()-timedelta(seconds=1); db.session.commit()
   criar_preview(FileStorage(stream=io.BytesIO(b'nome,cpf,email\nNovo,12345678908,novo@example.test\n'), filename='novo.csv'), 'clientes')
   db.session.expire_all()
   self.assertEqual(ImportPreview.query.filter(ImportPreview.expires_at < datetime.utcnow()).count(), 0)
 def test_invalid_internal_xml_returns_400_and_audits_are_redacted(self):
  wb=Workbook(); ws=wb.active; ws.title='Clientes'; ws.append(['nome','cpf','email']); ws.append(['Ana','12345678909','ana@example.test']); source=io.BytesIO(); wb.save(source)
  corrupted=io.BytesIO()
  with zipfile.ZipFile(io.BytesIO(source.getvalue())) as original, zipfile.ZipFile(corrupted, 'w') as target:
   for item in original.infolist(): target.writestr(item, b'<broken' if item.filename == 'xl/worksheets/sheet1.xml' else original.read(item.filename))
  bad=self.app.test_client().post('/api/v1/importacoes/preview',headers={'Authorization':'Bearer '+self._token(self.a)},data={'arquivo':(io.BytesIO(corrupted.getvalue()),'corrupt.xml.xlsx')},content_type='multipart/form-data'); self.assertEqual(bad.status_code,400)
  ok=self._preview(self.a,b'nome,cpf,email\nAuditada,12345678910,auditada@example.test\n'); self.assertEqual(ok.status_code,201)
  with self.app.app_context():
   entries=LogEntry.query.filter(LogEntry.acao.in_(['import_preview_rejected','import_preview_created'])).all(); self.assertTrue(entries)
   text=' '.join(str(entry.metadados) for entry in entries)
   self.assertNotIn('12345678909', text); self.assertNotIn('ana@example.test', text); self.assertNotIn('12345678910', text); self.assertNotIn('auditada@example.test', text)
   self.assertTrue(all({'usuarioId','contagens','resultado'}.issubset(entry.metadados) for entry in entries))
 def test_cli_purges_expired_previews_globally(self):
  with self.app.app_context():
   now=datetime.utcnow()-timedelta(seconds=1)
   db.session.add_all([ImportPreview(empresa_id=1,created_by_id=self.a,arquivo_hash='a'*64,plano={'clientes':[],'ucs':[],'usinas':[]},status='pronto',expires_at=now), ImportPreview(empresa_id=2,created_by_id=self.b,arquivo_hash='b'*64,plano={'clientes':[],'ucs':[],'usinas':[]},status='pronto',expires_at=now)])
   db.session.commit()
  result=self.app.test_cli_runner().invoke(args=['purge-import-previews'])
  self.assertEqual(result.exit_code,0,result.output)
  with self.app.app_context(): self.assertEqual(ImportPreview.query.filter(ImportPreview.expires_at < datetime.utcnow()).count(),0)
 def test_request_purge_does_not_remove_expired_preview_from_other_tenant(self):
  with self.app.app_context():
   expired=ImportPreview(empresa_id=2,created_by_id=self.b,arquivo_hash='c'*64,plano={'clientes':[],'ucs':[],'usinas':[]},status='pronto',expires_at=datetime.utcnow()-timedelta(seconds=1)); db.session.add(expired); db.session.commit(); other_id=expired.id
  response=self._preview(self.a,b'nome,cpf,email\nIsolada,12345678911,isolada@example.test\n'); self.assertEqual(response.status_code,201)
  with self.app.app_context(): self.assertIsNotNone(ImportPreview.query.filter_by(id=other_id,empresa_id=2).first())
 def test_rbac_and_tenant_ownership(self):
  self.assertEqual(self._preview(self.viewer,b'nome,cpf,email\nX,12345678904,x@x.test\n').status_code,403)
  r=self._preview(self.a,b'nome,cpf,email\nX,12345678904,x@x.test\n'); pid=r.json['data']['previewId']; self.assertEqual(self._commit(self.b,pid).status_code,404)
 def test_duplicate_preview_and_atomic_rollback(self):
  d=self._preview(self.a,b'nome,cpf,email\nX,12345678905,x@x.test\nX2,12345678905,y@x.test\n'); self.assertEqual(d.status_code,201); self.assertTrue(d.json['data']['erros'])
  # XLSX com cliente + UC que referencia CPF inexistente: commit deve reverter cliente recém-criado.
  wb=Workbook(); c=wb.active; c.title='Clientes'; c.append(['nome','cpf','email']); c.append(['Y','12345678906','y@x.test']); u=wb.create_sheet('UCs'); u.append(['clienteCpf','codigo']); u.append(['99999999999','UC-1']); out=io.BytesIO(); wb.save(out)
  r=self.app.test_client().post('/api/v1/importacoes/preview',headers={'Authorization':'Bearer '+self._token(self.a)},data={'arquivo':(io.BytesIO(out.getvalue()),'atomico.xlsx')},content_type='multipart/form-data'); pid=r.json['data']['previewId']; self.assertEqual(self._commit(self.a,pid).status_code,409)
  with self.app.app_context():
   self.assertIsNone(Client.query.filter_by(cpf='12345678906').first())
   self.assertEqual(LogEntry.query.filter_by(acao='import_confirm_failed', entidade_id=pid).one().metadados['resultado'],'falha')

if __name__=='__main__': unittest.main()
